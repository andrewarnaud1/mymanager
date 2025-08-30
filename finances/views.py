# finances/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator
from datetime import date, datetime, timedelta
from decimal import Decimal
import openpyxl
from accounts.decorators import employee_required

from django.contrib.auth.decorators import login_required

from .models import DailySale, MonthlySummary, ExcelImportLog
from .forms import DailySaleForm, ExcelImportForm, DateRangeFilterForm, MonthYearFilterForm

@employee_required
def dashboard_view(request):
    """
    Dashboard principal des finances avec KPIs et graphiques
    """
    # Période par défaut : mois actuel
    today = date.today()
    debut_mois = date(today.year, today.month, 1)
    
    # KPIs du mois actuel
    ventes_mois = DailySale.objects.filter(date__gte=debut_mois)
    ca_mois = ventes_mois.aggregate(total=Sum('total_journalier'))['total'] or 0
    nb_jours_mois = ventes_mois.filter(total_journalier__gt=0).count()
    ca_moyen_jour = ca_mois / nb_jours_mois if nb_jours_mois > 0 else 0
    
    # Évolution par rapport au mois précédent
    mois_precedent_debut = (debut_mois.replace(day=1) - timedelta(days=1)).replace(day=1)
    mois_precedent_fin = debut_mois - timedelta(days=1)
    ca_mois_precedent = DailySale.objects.filter(
        date__gte=mois_precedent_debut,
        date__lte=mois_precedent_fin
    ).aggregate(total=Sum('total_journalier'))['total'] or 0
    
    evolution = 0
    if ca_mois_precedent > 0:
        evolution = ((ca_mois - ca_mois_precedent) / ca_mois_precedent) * 100
    
    # Dernières ventes
    dernières_ventes = DailySale.objects.all().order_by('-date')[:7]
    
    # Données pour graphique (30 derniers jours)
    debut_graph = today - timedelta(days=30)
    ventes_graph = DailySale.objects.filter(
        date__gte=debut_graph
    ).order_by('date')
    
    # Écarts significatifs
    ecarts_significatifs = DailySale.objects.filter(
        date__gte=debut_mois
    ).filter(
        Q(ecart_total__lt=-1) | Q(ecart_total__gt=1)
    ).order_by('-date')
    
    # Résumés mensuels récents
    resumes_mensuels = MonthlySummary.objects.all()[:6]
    
    context = {
        'ca_mois': ca_mois,
        'ca_moyen_jour': ca_moyen_jour,
        'nb_jours_mois': nb_jours_mois,
        'evolution': evolution,
        'dernières_ventes': dernières_ventes,
        'ventes_graph': ventes_graph,
        'ecarts_significatifs': ecarts_significatifs,
        'resumes_mensuels': resumes_mensuels,
        'today': today,
    }
    
    return render(request, 'finances/dashboard.html', context)


def sales_list_view(request):
    """
    Liste des ventes avec filtres et pagination
    """
    ventes = DailySale.objects.all().order_by('-date')
    
    # Filtres
    filter_form = DateRangeFilterForm(request.GET)
    if filter_form.is_valid():
        date_debut = filter_form.cleaned_data.get('date_debut')
        date_fin = filter_form.cleaned_data.get('date_fin')
        
        if date_debut:
            ventes = ventes.filter(date__gte=date_debut)
        if date_fin:
            ventes = ventes.filter(date__lte=date_fin)
    
    # Recherche par commentaire
    search = request.GET.get('search')
    if search:
        ventes = ventes.filter(
            Q(commentaires__icontains=search) |
            Q(date__icontains=search)
        )
    
    # Statistiques de la sélection
    stats = ventes.aggregate(
        total_ca=Sum('total_journalier'),
        total_cb=Sum('cb_tpe'),
        total_especes=Sum('especes_reel'),
        total_tr=Sum('tr_reel'),
        total_ecarts=Sum('ecart_total'),
        nb_jours=Count('id')
    )
    
    # Pagination
    paginator = Paginator(ventes, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'filter_form': filter_form,
        'search': search,
        'stats': stats,
    }
    
    return render(request, 'finances/sales_list.html', context)


def sale_create_view(request):
    """
    Créer une nouvelle vente journalière
    """
    if request.method == 'POST':
        form = DailySaleForm(request.POST)
        if form.is_valid():
            sale = form.save()
            messages.success(request, f'Vente du {sale.date.strftime("%d/%m/%Y")} enregistrée avec succès !')
            return redirect('finances:sales_list')
    else:
        form = DailySaleForm()
    
    return render(request, 'finances/sale_form.html', {
        'form': form,
        'title': 'Ajouter une vente journalière'
    })


def sale_update_view(request, pk):
    """
    Modifier une vente journalière
    """
    sale = get_object_or_404(DailySale, pk=pk)
    
    if request.method == 'POST':
        form = DailySaleForm(request.POST, instance=sale)
        if form.is_valid():
            sale = form.save()
            messages.success(request, f'Vente du {sale.date.strftime("%d/%m/%Y")} mise à jour !')
            return redirect('finances:sales_list')
    else:
        form = DailySaleForm(instance=sale)
    
    return render(request, 'finances/sale_form.html', {
        'form': form,
        'title': f'Modifier la vente du {sale.date.strftime("%d/%m/%Y")}',
        'sale': sale
    })


def sale_delete_view(request, pk):
    """
    Supprimer une vente journalière
    """
    sale = get_object_or_404(DailySale, pk=pk)
    
    if request.method == 'POST':
        date_str = sale.date.strftime("%d/%m/%Y")
        sale.delete()
        messages.success(request, f'Vente du {date_str} supprimée.')
        return redirect('finances:sales_list')
    
    return render(request, 'finances/sale_confirm_delete.html', {'sale': sale})


def excel_import_view(request):
    """
    Import depuis fichier Excel
    """
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            overwrite = form.cleaned_data['overwrite_existing']
            
            try:
                result = import_excel_file(excel_file, overwrite)
                
                # Créer le log d'import
                log = ExcelImportLog.objects.create(
                    filename=excel_file.name,
                    nb_records_created=result['created'],
                    nb_records_updated=result['updated'],
                    nb_records_skipped=result['skipped'],
                    errors='\n'.join(result['errors']) if result['errors'] else '',
                    success=len(result['errors']) == 0
                )
                
                if result['errors']:
                    messages.warning(request, 
                        f"Import terminé avec {len(result['errors'])} erreurs. "
                        f"Créés: {result['created']}, Mis à jour: {result['updated']}, "
                        f"Ignorés: {result['skipped']}"
                    )
                else:
                    messages.success(request, 
                        f"Import réussi ! Créés: {result['created']}, "
                        f"Mis à jour: {result['updated']}, Ignorés: {result['skipped']}"
                    )
                
                return redirect('finances:dashboard')
                
            except Exception as e:
                ExcelImportLog.objects.create(
                    filename=excel_file.name,
                    errors=str(e),
                    success=False
                )
                messages.error(request, f"Erreur lors de l'import : {str(e)}")
    else:
        form = ExcelImportForm()
    
    # Historique des imports
    import_logs = ExcelImportLog.objects.all()[:10]
    
    return render(request, 'finances/excel_import.html', {
        'form': form,
        'import_logs': import_logs
    })


def monthly_reports_view(request):
    """
    Rapports mensuels avec comparaisons
    """
    filter_form = MonthYearFilterForm(request.GET)
    resumes = MonthlySummary.objects.all()
    
    if filter_form.is_valid():
        annee = filter_form.cleaned_data.get('annee')
        mois = filter_form.cleaned_data.get('mois')
        
        if annee:
            resumes = resumes.filter(annee=annee)
        if mois:
            resumes = resumes.filter(mois=int(mois))
    
    # Calculs pour graphiques année
    if filter_form.is_valid() and filter_form.cleaned_data.get('annee'):
        annee = filter_form.cleaned_data['annee']
        donnees_annee = []
        for mois in range(1, 13):
            try:
                resume = MonthlySummary.objects.get(annee=annee, mois=mois)
                donnees_annee.append({
                    'mois': mois,
                    'ca': float(resume.total_ca),
                    'jours': resume.jours_ouverture
                })
            except MonthlySummary.DoesNotExist:
                donnees_annee.append({
                    'mois': mois,
                    'ca': 0,
                    'jours': 0
                })
    else:
        donnees_annee = []
    
    context = {
        'resumes': resumes,
        'filter_form': filter_form,
        'donnees_annee': donnees_annee,
    }
    
    return render(request, 'finances/monthly_reports.html', context)


def import_excel_file(excel_file, overwrite=False):
    """
    Fonction utilitaire pour importer un fichier Excel
    """
    result = {
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': []
    }
    
    try:
        # Lire le fichier Excel
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Ignorer les feuilles vides
            if worksheet.max_row < 2:
                continue
            
            # Lire les en-têtes (première ligne)
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)
            
            # Vérifier que c'est bien le bon format
            if 'DATE' not in headers:
                continue
            
            # Mapper les colonnes
            col_mapping = {}
            for idx, header in enumerate(headers):
                if header == 'DATE':
                    col_mapping['date'] = idx
                elif header == 'CB TPE':
                    col_mapping['cb_tpe'] = idx
                elif header == 'CB CAISSE':
                    col_mapping['cb_caisse'] = idx
                elif header == 'ESPECES REEL':
                    col_mapping['especes_reel'] = idx
                elif header == 'ESPECES CAISSE':
                    col_mapping['especes_caisse'] = idx
                elif header == 'TICKETS RESTAU REEL':
                    col_mapping['tr_reel'] = idx
                elif header == 'TICKETS RESTAU CAISSE':
                    col_mapping['tr_caisse'] = idx
                elif header == 'TOTAL':
                    col_mapping['total'] = idx
                elif header == 'Commentaires':
                    col_mapping['commentaires'] = idx
                elif header in ['NOMBRE DE CLIENTS', 'NOMBRE DE CLIENTS ']:
                    col_mapping['nombre_clients'] = idx
                elif header in ['TICKET MOYEN', 'TICKET MOYEN PAR CLIENTS', 'TICKET MOYEN PAR CLEINT']:
                    col_mapping['ticket_moyen'] = idx
            
            # Traiter chaque ligne de données
            for row_num in range(2, worksheet.max_row + 1):
                try:
                    row = worksheet[row_num]
                    
                    # Extraire la date
                    if 'date' not in col_mapping:
                        continue
                    
                    date_cell = row[col_mapping['date']].value
                    if not date_cell:
                        continue
                    
                    # Convertir la date
                    if isinstance(date_cell, datetime):
                        sale_date = date_cell.date()
                    elif isinstance(date_cell, str):
                        # Essayer de parser différents formats de date
                        try:
                            if 'janvier' in date_cell.lower() or 'février' in date_cell.lower():
                                # Format français "mardi 1 avril 2025"
                                continue  # Ignorer les formats trop complexes pour le moment
                            else:
                                sale_date = datetime.strptime(date_cell, '%Y-%m-%d').date()
                        except:
                            continue
                    else:
                        continue
                    
                    # Vérifier si l'enregistrement existe déjà
                    try:
                        existing_sale = DailySale.objects.get(date=sale_date)
                        if not overwrite:
                            result['skipped'] += 1
                            continue
                        sale = existing_sale
                        is_update = True
                    except DailySale.DoesNotExist:
                        sale = DailySale(date=sale_date)
                        is_update = False
                    
                    # Extraire les données
                    def get_decimal_value(col_key):
                        if col_key not in col_mapping:
                            return None
                        value = row[col_mapping[col_key]].value
                        if value is None or value == '':
                            return None
                        try:
                            return Decimal(str(value))
                        except:
                            return None
                    
                    def get_int_value(col_key):
                        if col_key not in col_mapping:
                            return None
                        value = row[col_mapping[col_key]].value
                        if value is None or value == '':
                            return None
                        try:
                            return int(float(value))
                        except:
                            return None
                    
                    def get_string_value(col_key):
                        if col_key not in col_mapping:
                            return ''
                        value = row[col_mapping[col_key]].value
                        return str(value) if value is not None else ''
                    
                    # Assigner les valeurs
                    sale.cb_caisse = get_decimal_value('cb_caisse')
                    sale.cb_tpe = get_decimal_value('cb_tpe')
                    sale.especes_caisse = get_decimal_value('especes_caisse')
                    sale.especes_reel = get_decimal_value('especes_reel')
                    sale.tr_caisse = get_decimal_value('tr_caisse')
                    sale.tr_reel = get_decimal_value('tr_reel')
                    sale.nombre_clients = get_int_value('nombre_clients')
                    sale.commentaires = get_string_value('commentaires')
                    sale.imported_from_excel = True
                    
                    # Sauvegarder
                    sale.save()
                    
                    if is_update:
                        result['updated'] += 1
                    else:
                        result['created'] += 1
                
                except Exception as e:
                    result['errors'].append(f"Ligne {row_num} de {sheet_name}: {str(e)}")
        
        workbook.close()
        
    except Exception as e:
        result['errors'].append(f"Erreur générale: {str(e)}")
    
    return result

@login_required
def api_sales_data(request):
    """
    API pour récupérer les données de ventes (pour graphiques)
    """
    # Paramètres
    days = int(request.GET.get('days', 30))
    end_date = date.today()
    start_date = end_date - timedelta(days=days)
    
    # Récupérer les données
    sales = DailySale.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).order_by('date')
    
    data = []
    for sale in sales:
        data.append({
            'date': sale.date.strftime('%Y-%m-%d'),
            'total': float(sale.total_journalier),
            'cb': float(sale.chiffre_affaires_cb),
            'especes': float(sale.chiffre_affaires_especes),
            'tr': float(sale.chiffre_affaires_tr),
            'ecart': float(sale.ecart_total)
        })
    
    return JsonResponse({'data': data})