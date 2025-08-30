# staff/utils.py
from django.http import HttpResponse
from datetime import datetime, timedelta
import io

# Imports pour PDF (ReportLab)
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Imports pour Excel (openpyxl)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_schedule_pdf(schedule, employee=None):
    """Génère un PDF du planning hebdomadaire"""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("ReportLab n'est pas installé.")
    
    filename = f"planning_{schedule.week_start.strftime('%Y%m%d')}"
    if employee:
        filename += f"_{employee.last_name}_{employee.first_name}"
    filename += ".pdf"
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1,
        spaceAfter=0.5*cm
    )
    
    elements = []
    
    # Titre
    if employee:
        title_text = f"Planning de {employee.full_name}<br/>{schedule.week_range_display}"
    else:
        title_text = f"Planning d'équipe<br/>{schedule.week_range_display}"
    
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Préparer les données
    if employee:
        data = _prepare_individual_schedule_data(schedule, employee)
    else:
        data = _prepare_team_schedule_data(schedule)
    
    # Créer le tableau
    table = Table(data)
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


def generate_schedule_excel(schedule, employee=None):
    """Génère un fichier Excel du planning"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl n'est pas installé.")
    
    wb = Workbook()
    ws = wb.active
    
    if employee:
        ws.title = f"{employee.last_name} {employee.first_name}"
        title = f"Planning de {employee.full_name} - {schedule.week_range_display}"
    else:
        ws.title = "Planning équipe"
        title = f"Planning d'équipe - {schedule.week_range_display}"
    
    # Titre
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    if employee:
        _write_individual_excel_data(ws, schedule, employee, 3)
    else:
        _write_team_excel_data(ws, schedule, 3)
    
    # Ajuster largeurs
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    filename = f"planning_{schedule.week_start.strftime('%Y%m%d')}"
    if employee:
        filename += f"_{employee.last_name}_{employee.first_name}"
    filename += ".xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    buffer = io.BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    buffer.close()
    
    return response


def _prepare_individual_schedule_data(schedule, employee):
    """Prépare les données pour le planning individuel PDF"""
    data = [['Jour', 'Date', 'Horaires', 'Durée', 'Notes']]
    
    day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
    for i in range(7):
        current_date = schedule.week_start + timedelta(days=i)
        day_name = day_names[i]
        
        day_shifts = schedule.shifts.filter(
            employee=employee,
            date=current_date
        ).order_by('start_time')
        
        if day_shifts:
            for j, shift in enumerate(day_shifts):
                row_data = [
                    day_name if j == 0 else '',
                    current_date.strftime('%d/%m') if j == 0 else '',
                    f"{shift.start_time.strftime('%H:%M')} - {shift.end_time.strftime('%H:%M')}",
                    shift.duration_display,
                    shift.notes or '-'
                ]
                data.append(row_data)
        else:
            data.append([day_name, current_date.strftime('%d/%m'), 'Repos', '-', '-'])
    
    return data


def _prepare_team_schedule_data(schedule):
    """Prépare les données pour le planning d'équipe PDF"""
    day_names = ['Employé', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    data = [day_names]
    
    employees = set()
    for shift in schedule.shifts.all():
        employees.add(shift.employee)
    
    employees = sorted(employees, key=lambda e: (e.last_name, e.first_name))
    
    for employee in employees:
        row = [employee.full_name]
        
        for i in range(7):
            current_date = schedule.week_start + timedelta(days=i)
            day_shifts = schedule.shifts.filter(
                employee=employee,
                date=current_date
            ).order_by('start_time')
            
            if day_shifts:
                shifts_text = []
                for shift in day_shifts:
                    shifts_text.append(
                        f"{shift.start_time.strftime('%H:%M')}-{shift.end_time.strftime('%H:%M')}"
                    )
                row.append('\n'.join(shifts_text))
            else:
                row.append('-')
        
        data.append(row)
    
    return data


def _write_individual_excel_data(ws, schedule, employee, start_row):
    """Écrit les données du planning individuel dans Excel"""
    headers = ['Jour', 'Date', 'Horaires', 'Durée', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = start_row + 1
    day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
    for i in range(7):
        current_date = schedule.week_start + timedelta(days=i)
        day_name = day_names[i]
        
        day_shifts = schedule.shifts.filter(
            employee=employee,
            date=current_date
        ).order_by('start_time')
        
        if day_shifts:
            for j, shift in enumerate(day_shifts):
                ws.cell(row=row, column=1, value=day_name if j == 0 else '')
                ws.cell(row=row, column=2, value=current_date.strftime('%d/%m') if j == 0 else '')
                ws.cell(row=row, column=3, value=f"{shift.start_time.strftime('%H:%M')} - {shift.end_time.strftime('%H:%M')}")
                ws.cell(row=row, column=4, value=shift.duration_display)
                ws.cell(row=row, column=5, value=shift.notes or '-')
                row += 1
        else:
            ws.cell(row=row, column=1, value=day_name)
            ws.cell(row=row, column=2, value=current_date.strftime('%d/%m'))
            ws.cell(row=row, column=3, value='Repos')
            ws.cell(row=row, column=4, value='-')
            ws.cell(row=row, column=5, value='-')
            row += 1


def _write_team_excel_data(ws, schedule, start_row):
    """Écrit les données du planning d'équipe dans Excel"""
    day_headers = ['Employé', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
    for col, header in enumerate(day_headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = start_row + 1
    employees = set()
    for shift in schedule.shifts.all():
        employees.add(shift.employee)
    
    employees = sorted(employees, key=lambda e: (e.last_name, e.first_name))
    
    for employee in employees:
        ws.cell(row=row, column=1, value=employee.full_name)
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        for i in range(7):
            current_date = schedule.week_start + timedelta(days=i)
            day_shifts = schedule.shifts.filter(
                employee=employee,
                date=current_date
            ).order_by('start_time')
            
            col = i + 2
            if day_shifts:
                shifts_text = []
                for shift in day_shifts:
                    shifts_text.append(
                        f"{shift.start_time.strftime('%H:%M')}-{shift.end_time.strftime('%H:%M')}"
                    )
                ws.cell(row=row, column=col, value='\n'.join(shifts_text))
            else:
                ws.cell(row=row, column=col, value='-')
        
        row += 1